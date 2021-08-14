import openpyxl as pyxl

path = "C:\\Users\\Sayed\\Documents\\GitHub\\Python3\\TheWorld\\Country_Capital_Currency_population_continent.xlsx"

wb_obj = pyxl.load_workbook(path)
sheet_obj = wb_obj.active

maxRow = sheet_obj.max_row
maxClm = sheet_obj.max_column
country_list = []
country_details = {}
details = []
for i in range(2, maxRow + 1):
    cell_obj = sheet_obj.cell(row = i, column = 2)
    country_list.append(cell_obj.value)
    temp_cuntry = cell_obj.value
    for x in range(3,maxClm+1):
        cell_obj = sheet_obj.cell(row = i, column = x)
        details.append(cell_obj.value)

    #print(details)
    country_details[temp_cuntry] = details
    details = []


a = []
print(country_details)
print("------------------------------------------------------------------------")
print("------------------------------------------------------------------------")
print(country_list)
print("------------------------------------------------------------------------")
print("------------------------------------------------------------------------")
a = country_details['Venezuela']
print(a)
