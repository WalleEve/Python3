from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
import openpyxl as pyxl


root = Tk()
root.title("Country Details")
root.geometry("400x300")

# Global Varibale
global vSearch
global noticeLable
global searchCombo
global country_details
global country_list

# Fetch country name
def fetchCuntryExl():
    #global country_list
    #global country_details

    path = "C:\\Users\\Sayed\\Documents\\GitHub\\Python3\\TheWorld\\Country_Capital_Currency_population_continent.xlsx"

    wb_obj = pyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    maxRow = sheet_obj.max_row
    maxClm = sheet_obj.max_column
    country_list = []
    country_details = {}
    details = []
    for i in range(2, maxRow + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        country_list.append(cell_obj.value)
        temp_cuntry = cell_obj.value
        for x in range(3,maxClm+1):
            cell_obj = sheet_obj.cell(row = i, column = x)
            details.append(cell_obj.value)

        #print(details)
        country_details[temp_cuntry] = details
        details = []
    return country_list, country_details



# get the flag
def getflag(cName):
    image1 = Image.open("all_flags\\"+cName.lower()+'.gif') # "India_Flag.jpg"
    image1 = image1.resize((150, 85))
    test = ImageTk.PhotoImage(image1)
    return test

def clearscr():
    searchCombo.set("")
    noticeLable.configure(text="")



# Functions:
def search_detail():
    global temptSearch
    vDetails = []


    search_dashboard = Toplevel(root)
    search_dashboard.title("Dountry Dashboard")
    search_dashboard.geometry("600x400+200+200")

    tSearch = vSearch.get()
    if ' ' in tSearch:
        temptSearch = tSearch.replace(' ','_').replace('-','_')

    else:
        temptSearch = tSearch
    # Clear Search Screen
    clearscr()
    # Country Details from Dictionary Object
    vDetails = country_details[tSearch]

    # Country Frame
    cFrame = Frame(search_dashboard)
    cFrame.place(x=5, y=5, width=400, height=90)

    cLable = Label(cFrame, text=tSearch, font=("Calibri", 15, "bold"))
    cLable.place(x=40, y=30)

    # Flag Frame
    fFrame = Frame(search_dashboard)
    fFrame.place(x=410, y=5, width=150, height=90)

    img = getflag(temptSearch)
    label1 = Label(fFrame, image=img)
    label1.image = img
    label1.place(x=2, y=5)

    # Detail Frame
    dFrame = Frame(search_dashboard, bg="white")
    dFrame.place(x=5, y=100, width=585, height=290)

    # Label:
    capital = Label(dFrame, text="Capital: ", font=("Calibri", 15, "bold"), bg="white")
    capital.grid(row=0, column=0, sticky=W, padx=5, pady=5)
    currency = Label(dFrame, text="Currency: ", font=("Calibri", 15, "bold"), bg="white")
    currency.grid(row=1, column=0, sticky=W, padx=5, pady=5)
    language = Label(dFrame, text="Language: ", font=("Calibri", 15, "bold"), bg="white")
    language.grid(row=2, column=0, sticky=W, padx=5, pady=5)
    population = Label(dFrame, text="Population: ", font=("Calibri", 15, "bold"), bg="white")
    population.grid(row=3, column=0, sticky=W, padx=5, pady=5)

    print(vDetails)
    Label(dFrame, text=vDetails[1], bg="white", font=("Calibri", 12, "bold") ).grid(row=0, column=1, sticky=W, padx=5, pady=5)
    Label(dFrame, text=vDetails[2], bg="white", font=("Calibri", 12, "bold") ).grid(row=1, column=1, sticky=W, padx=5, pady=5)
    Label(dFrame, text="Language", bg="white", font=("Calibri", 12, "bold") ).grid(row=2, column=1, sticky=W, padx=5, pady=5) #  vDetails[1]
    Label(dFrame, text=vDetails[0], bg="white", font=("Calibri", 12, "bold") ).grid(row=3, column=1, sticky=W, padx=5, pady=5)
    Label(dFrame, text="", bg="white", font=("Calibri", 12, "bold") ).grid(row=4, column=0, sticky=W, padx=5, pady=5)
    Label(dFrame, text="", bg="white", font=("Calibri", 12, "bold") ).grid(row=5, column=0, sticky=W, padx=5, pady=5)



    # button
    exitButton = Button(dFrame, text="Exit", font=("Calibri", 12, "bold"), height = 1, width=12)
    exitButton.grid(row=6, column=0, columnspan=1, sticky=W)
    wikiButton = Button(dFrame, text="Wiki", font=("Calibri", 12, "bold"), height = 1, width=12)
    wikiButton.grid(row=6, column=1, columnspan=1, sticky=W)


def searchExp():
    noticeLable.configure(text="*Please choose one option!")

def search():
    tSearch = vSearch.get()
    if tSearch == "Pick an Option" or tSearch == '':
        searchExp()
    else:
        search_detail()
# Frame:
vSearch = StringVar()
tileFrame = Frame(root, bg="aquamarine1")
tileFrame.place(x=5, y=5, width=390, height=40)

mainFrame = Frame(root, bg="aqua")
mainFrame.place(x=5, y=50, width=390, height=242)

tileLabel = Label(tileFrame, text="Country Details", bg="aquamarine1", font=("Calibri", 15, "bold"))
tileLabel.pack()

searchBtn = Button(mainFrame, text="Search", font=("Calibri", 12, "bold"), command=search,  height = 1, width=12)
searchBtn.pack(side=LEFT,padx=10)

# COMBO BOX
#vlist = ["India", "China", "Japan","Nepal", "Bhutan", "Saint Vincent and the Grenadines","Micronesia","Sao Tome and Principe","East Timor (Timor-Leste)"]
country_list, country_details  = fetchCuntryExl()

def check(e):
    v= vSearch.get()
    if v=='':
        data= country_list
    else:
        data=[]
        for item in country_list:
            if v.lower() in item.lower():
                data.append(item)
    #update(data)

def update(data):
   # Clear the Combobox
   menu.delete(0, END)
   # Add values to the combobox
   for value in data:
      menu.insert(END,value)


searchCombo = ttk.Combobox(mainFrame, values = country_list, font=("Calibri", 12, "bold"), textvariable = vSearch, height = 2, width=22)
searchCombo.set("Pick an Option")
searchCombo.pack(side=LEFT, padx = 5, pady = 5)
searchCombo.bind('<KeyRelease>',check)

# Notice
noticeFrame = Frame(mainFrame, bg="aqua")
noticeFrame.place(x=75, y=150, width=200, height=30)
noticeLable = Label(noticeFrame, text="",fg="red", bg="aqua", font=("Calibri", 10)) #
noticeLable.pack(side=LEFT, fill=BOTH, padx=15, pady=5 )





root.mainloop()
